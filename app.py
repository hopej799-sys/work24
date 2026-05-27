import streamlit as st
import requests
import xml.etree.ElementTree as ET
import pandas as pd
import io
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from supabase import create_client

st.set_page_config(
    page_title="연계채용정보 모니터링",
    page_icon="🔍",
    layout="wide",
)

API_URL = "https://www.work24.go.kr/cm/openApi/call/wk/callOpenApiSvcInfo220L01.do"

COLUMNS = [
    ("sysGbnNm",      "오류구분"),
    ("iorgGbnm",      "연계기관명"),
    ("wantedAuthNo",  "공고번호"),
    ("createDtm",     "모니터링일시"),
    ("ifDtm",         "연계일시"),
    ("errCont",       "에러내용"),
    ("lawVoltDobtYn", "법위반의심 여부"),
    ("lawMappCont",   "법령 맵핑 내용"),
]

STATUS_OPTIONS = ["미검토", "이상없음", "게재중단"]


@st.cache_resource
def _sb():
    return create_client(st.secrets["supabase"]["url"], st.secrets["supabase"]["key"])


@st.cache_data(ttl=5)
def load_store() -> dict:
    rows = _sb().table("memo_store").select("wanted_auth_no, status, memo").execute().data
    return {r["wanted_auth_no"]: {"처리상태": r["status"], "메모": r["memo"]} for r in rows}


def save_store(store: dict):
    sb = _sb()
    to_delete = set(load_store().keys()) - set(store.keys())
    if to_delete:
        sb.table("memo_store").delete().in_("wanted_auth_no", list(to_delete)).execute()
    if store:
        rows = [{"wanted_auth_no": k, "status": v["처리상태"], "memo": v["메모"]} for k, v in store.items()]
        sb.table("memo_store").upsert(rows).execute()
    load_store.clear()


def build_store(base_df: pd.DataFrame, visible_wanted: set, edited: pd.DataFrame) -> dict:
    """현재 화면(edited) + 필터로 숨겨진 행(base_df)을 합쳐 저장할 dict 생성"""
    store = {}
    # 필터로 숨겨진 행은 base_df 값 사용
    for _, row in base_df[~base_df["공고번호"].isin(visible_wanted)].iterrows():
        wanted = row["공고번호"]
        status = row["처리상태"]
        memo   = str(row["메모"] or "")
        if wanted and (status != "미검토" or memo.strip()):
            store[wanted] = {"처리상태": status, "메모": memo}
    # 현재 화면 행은 edited 값 사용 (data_editor의 실제 현재 값)
    for i in range(len(edited)):
        wanted = edited.iloc[i]["공고번호"]
        status = edited.iloc[i]["처리상태"]
        raw    = edited.iloc[i]["메모"]
        memo   = "" if pd.isna(raw) else str(raw)
        if not wanted:
            continue
        if status != "미검토" or memo.strip():
            store[wanted] = {"처리상태": status, "메모": memo}
        elif wanted in store:
            del store[wanted]
    return store


# ── 사이드바 ───────────────────────────────────
with st.sidebar:
    st.title("🔍 조회 조건")

    auth_key = st.text_input("인증키 *", type="password", placeholder="발급받은 인증키 입력")

    st.markdown("**조회 기간 *** (최대 3일)")
    col1, col2 = st.columns(2)
    with col1:
        start_date = st.date_input("시작일", value=date.today() - timedelta(days=2))
    with col2:
        end_date = st.date_input("종료일", value=date.today())

    wanted_auth_no = st.text_input("공고번호 (선택)", placeholder="예) K123456789")
    law_volt = st.selectbox("법위반의심 여부 (선택)", ["전체", "Y (의심 항목만)"])

    st.divider()
    st.markdown("**조회 필터**")
    exclude_employment = st.checkbox(
        "고용형태 에러 제외",
        help="에러내용에 '고용형태'가 포함된 항목을 조회 결과에서 제외합니다",
    )

    search_btn = st.button("🔎 조회", use_container_width=True, type="primary")


# ── 유효성 검사 ────────────────────────────────
def validate():
    if not auth_key:
        st.error("인증키를 입력해 주세요.")
        return False
    delta = (end_date - start_date).days
    if delta < 0:
        st.error("종료일이 시작일보다 빠릅니다.")
        return False
    if delta > 2:
        st.error("최대 3일 범위까지만 조회 가능합니다.")
        return False
    return True


# ── API 호출 ───────────────────────────────────
def fetch(start, end, auth, wanted_no, law_y):
    params = {
        "authKey":    auth,
        "returnType": "XML",
        "callTp":     "D",
        "ifDtmStdt":  start.strftime("%Y%m%d"),
        "ifDtmEndt":  end.strftime("%Y%m%d"),
    }
    if wanted_no:
        params["wantedAuthNo"] = wanted_no
    if law_y == "Y (의심 항목만)":
        params["lawVoltDobtYn"] = "Y"

    resp = requests.get(API_URL, params=params, timeout=30)
    resp.raise_for_status()
    return resp.text


# ── XML 파싱 ───────────────────────────────────
def parse(xml_text):
    root = ET.fromstring(xml_text)
    rows = []
    for item in root.iter("monitoringErrInfo"):
        row = {}
        for key, label in COLUMNS:
            node = item.find(key)
            row[label] = node.text.strip() if node is not None and node.text else ""
        rows.append(row)
    df = pd.DataFrame(rows, columns=[label for _, label in COLUMNS])
    df.insert(0, "처리상태", "미검토")
    df.insert(1, "메모", "")
    return df


# ── 엑셀 생성 ──────────────────────────────────
def make_excel(df, start, end):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "모니터링결과"

    header_font  = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
    header_fill  = PatternFill("solid", start_color="1F5C8B")
    cell_font    = Font(name="맑은 고딕", size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin         = Side(style="thin", color="CCCCCC")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill     = PatternFill("solid", start_color="EAF2FB")
    green_fill   = PatternFill("solid", start_color="D9F2D9")
    red_fill     = PatternFill("solid", start_color="FFD9D9")

    ncols = len(df.columns)
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    t = ws["A1"]
    t.value     = f"연계채용정보 모니터링 결과  |  {start.strftime('%Y%m%d')} ~ {end.strftime('%Y%m%d')}"
    t.font      = Font(name="맑은 고딕", bold=True, size=12, color="1F5C8B")
    t.alignment = center_align
    ws.row_dimensions[1].height = 28

    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(row=2, column=ci, value=col)
        c.font = header_font; c.fill = header_fill
        c.alignment = center_align; c.border = border
    ws.row_dimensions[2].height = 22

    long_cols = {"에러내용", "법령 맵핑 내용"}
    for ri, row in enumerate(df.itertuples(index=False), 3):
        status = row[0]
        if status == "이상없음":
            row_fill = green_fill
        elif status == "게재중단":
            row_fill = red_fill
        else:
            row_fill = alt_fill if ri % 2 == 1 else None

        for ci, (col, val) in enumerate(zip(df.columns, row), 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = cell_font; c.border = border
            c.alignment = left_align if col in long_cols else center_align
            if row_fill:
                c.fill = row_fill
        ws.row_dimensions[ri].height = 18

    col_widths = [12, 30, 18, 20, 20, 20, 20, 50, 16, 40]
    for i, w in enumerate(col_widths[:ncols], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── 메인 ───────────────────────────────────────
st.title("📋 연계채용정보 모니터링 결과조회")
st.caption("한국고용정보원 Work24 Open API")

if search_btn:
    if not validate():
        st.stop()

    with st.spinner("데이터 조회 중..."):
        try:
            xml_text = fetch(start_date, end_date, auth_key, wanted_auth_no, law_volt)
        except requests.HTTPError as e:
            st.error(f"API 오류: {e}")
            st.stop()
        except Exception as e:
            st.error(f"오류 발생: {e}")
            st.stop()

    try:
        df = parse(xml_text)
    except ET.ParseError as e:
        st.error(f"XML 파싱 오류: {e}")
        st.stop()

    if exclude_employment:
        df = df[~df["에러내용"].str.contains("고용형태", na=False)].reset_index(drop=True)

    # Supabase 저장값을 base_df에 반영 (이후 base_df는 수정하지 않음)
    stored = load_store()
    for idx in df.index:
        key = df.at[idx, "공고번호"]
        if key and key in stored:
            df.at[idx, "처리상태"] = stored[key]["처리상태"]
            df.at[idx, "메모"]     = stored[key]["메모"]

    st.session_state["base_df"] = df
    st.session_state["period"]  = (start_date, end_date)
    st.session_state.pop("data_editor", None)  # 이전 편집 diff 초기화


if "base_df" not in st.session_state:
    st.info("← 왼쪽 사이드바에서 조회 조건을 입력하고 **조회** 버튼을 누르세요.")
    st.stop()

base_df = st.session_state["base_df"]
start_saved, end_saved = st.session_state["period"]

if base_df.empty:
    st.info("조회된 데이터가 없습니다.")
    st.stop()

# 요약 지표 (base_df 기준)
c1, c2, c3, c4 = st.columns(4)
c1.metric("총 건수",   f"{len(base_df):,}건")
c2.metric("법위반의심", f"{(base_df['법위반의심 여부'] == 'Y').sum():,}건")
c3.metric("미검토",    f"{(base_df['처리상태'] == '미검토').sum():,}건")
c4.metric("처리완료",  f"{(base_df['처리상태'] != '미검토').sum():,}건")

st.divider()

# 결과 필터
with st.expander("🔧 결과 필터", expanded=True):
    f1, f2, f3 = st.columns([2, 2, 1])
    with f1:
        status_filter = st.selectbox("처리 상태", ["전체", "미검토", "이상없음", "게재중단"])
    with f2:
        err_type_filter = st.selectbox("오류구분", ["전체", "사전필터링", "구인 모니터링"])
    with f3:
        st.markdown("<br>", unsafe_allow_html=True)
        law_only  = st.checkbox("법령 맵핑 내용 있는것만")
        memo_only = st.checkbox("메모 있는것만")

# 필터 적용
filtered = base_df.copy()
if status_filter != "전체":
    filtered = filtered[filtered["처리상태"] == status_filter]
if err_type_filter == "사전필터링":
    filtered = filtered[filtered["오류구분"].str.contains("사전", na=False)]
elif err_type_filter == "구인 모니터링":
    filtered = filtered[filtered["오류구분"].str.contains("구인", na=False)]
if law_only:
    filtered = filtered[filtered["법령 맵핑 내용"].str.strip().ne("")]
if memo_only:
    filtered = filtered[filtered["메모"].str.strip().ne("")]

filtered_display  = filtered.reset_index(drop=True)
visible_wanted    = set(filtered_display["공고번호"])

st.caption(f"필터 결과: {len(filtered_display):,}건 / 전체 {len(base_df):,}건")

edited = st.data_editor(
    filtered_display,
    column_config={
        "처리상태": st.column_config.SelectboxColumn(
            "처리상태",
            options=STATUS_OPTIONS,
            required=True,
            width="small",
        ),
        "메모":           st.column_config.TextColumn("메모",           width="medium"),
        "에러내용":       st.column_config.TextColumn("에러내용",       width="large"),
        "법령 맵핑 내용": st.column_config.TextColumn("법령 맵핑 내용", width="large"),
    },
    hide_index=True,
    use_container_width=True,
    height=450,
    key="data_editor",
)

st.divider()

# 저장 여부 판단
current_store = build_store(base_df, visible_wanted, edited)
has_changes   = current_store != load_store()

btn_col, dl_col = st.columns([1, 2])

with btn_col:
    save_label = "💾 저장" + (" ●" if has_changes else "")
    if st.button(save_label, type="primary", use_container_width=True, key="save_btn"):
        try:
            save_store(current_store)
        except Exception as e:
            st.error(f"저장 실패: {e}")
            st.stop()

        # base_df에 저장된 값 반영 후 diff 초기화
        new_base = base_df.copy()
        for idx in new_base.index:
            key = new_base.at[idx, "공고번호"]
            if key and key in current_store:
                new_base.at[idx, "처리상태"] = current_store[key]["처리상태"]
                new_base.at[idx, "메모"]     = current_store[key]["메모"]
            elif key:
                new_base.at[idx, "처리상태"] = "미검토"
                new_base.at[idx, "메모"]     = ""
        st.session_state["base_df"] = new_base
        st.session_state.pop("data_editor", None)

        st.toast(f"저장 완료 ({len(current_store)}건)", icon="✅")
        st.rerun()

with dl_col:
    # 엑셀 = base_df에 current_store 반영한 전체 데이터
    export_df = base_df.copy()
    for idx in export_df.index:
        key = export_df.at[idx, "공고번호"]
        if key and key in current_store:
            export_df.at[idx, "처리상태"] = current_store[key]["처리상태"]
            export_df.at[idx, "메모"]     = current_store[key]["메모"]

    excel_buf = make_excel(export_df, start_saved, end_saved)
    filename  = f"모니터링결과_{start_saved.strftime('%Y%m%d')}-{end_saved.strftime('%Y%m%d')}.xlsx"
    st.download_button(
        label="📥 엑셀 다운로드 (전체)",
        data=excel_buf,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
