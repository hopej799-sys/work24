import streamlit as st
import requests
import xml.etree.ElementTree as ET
import pandas as pd
import io
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from supabase import create_client

st.set_page_config(
    page_title="연계채용정보 모니터링",
    page_icon="🔍",
    layout="wide",
)

API_URL = "https://www.work24.go.kr/cm/openApi/call/wk/callOpenApiSvcInfo220L01.do"

COLUMNS = [
    ("sysGbnNm",      "오류구분"),
    ("wantedAuthNo",  "공고번호"),
    ("createDtm",     "모니터링일시"),
    ("ifDtm",         "연계일시"),
    ("errCont",       "에러내용"),
    ("wantedInfoUrl", "구인공고 URL"),
    ("lawVoltDobtYn", "법위반의심 여부"),
    ("lawMappCont",   "법령 맵핑 내용"),
]

STATUS_OPTIONS = ["미검토", "검토중", "검토완료", "이상없음", "게재중단"]


@st.cache_resource
def _sb():
    return create_client(st.secrets["supabase"]["url"], st.secrets["supabase"]["key"])


@st.cache_data(ttl=5)
def load_store() -> dict:
    rows = _sb().table("memo_store").select("wanted_auth_no, status, memo, status_changed_at").execute().data
    return {
        r["wanted_auth_no"]: {
            "처리상태": r["status"],
            "메모": r["memo"],
            "상태변경일": r.get("status_changed_at") or "",
        }
        for r in rows
    }


def save_store(store: dict, to_delete: set = None):
    sb = _sb()
    if to_delete:
        sb.table("memo_store").delete().in_("wanted_auth_no", list(to_delete)).execute()
    if store:
        rows = [{"wanted_auth_no": k, "status": v["처리상태"], "memo": v["메모"],
                 "status_changed_at": v.get("상태변경일", "")}
                for k, v in store.items()]
        sb.table("memo_store").upsert(rows).execute()
    load_store.clear()


def _get_pending_from_editor():
    """저장 시점에 data_editor 위젯 상태에서 편집 내역을 계산."""
    fd = st.session_state.get("_fd", pd.DataFrame())
    editor = st.session_state.get("data_editor", {})
    result = {}
    for str_idx, changes in editor.get("edited_rows", {}).items():
        row_idx = int(str_idx)
        if row_idx >= len(fd):
            continue
        k = fd.iloc[row_idx]["공고번호"]
        if not k:
            continue
        base_s = fd.iloc[row_idx]["처리상태"]
        base_m = str(fd.iloc[row_idx]["메모"] or "")
        base_c = str(fd.iloc[row_idx].get("상태변경일", "") or "")
        new_s  = changes.get("처리상태", base_s)
        new_m  = str(changes.get("메모", base_m) or "")
        if new_s != base_s:
            changed_at = date.today().strftime("%Y-%m-%d") if new_s != "미검토" else ""
        else:
            changed_at = base_c
        result[k] = {"처리상태": new_s, "메모": new_m, "상태변경일": changed_at}
    return result


# ── 사이드바 ───────────────────────────────────
with st.sidebar:
    st.title("🔍 조회 조건")

    auth_key = st.text_input("인증키 *", type="password",
                              placeholder="발급받은 인증키 입력")

    st.markdown("**조회 기간 *** (최대 31일)")
    col1, col2 = st.columns(2)
    with col1:
        start_date = st.date_input("시작일", value=date.today() - timedelta(days=2))
    with col2:
        end_date = st.date_input("종료일", value=date.today())

    wanted_auth_no = st.text_input("공고번호 (선택)", placeholder="예) K123456789")

    st.divider()
    st.markdown("**조회 필터**")
    exclude_employment = st.checkbox(
        "고용형태 에러 제외",
        help="에러내용에 '고용형태'가 포함된 항목을 조회 결과에서 제외합니다",
    )

    search_btn = st.button("🔎 조회", use_container_width=True, type="primary")

    st.divider()
    st.markdown("**관리**")
    st.page_link("pages/monthly.py", label="📊 월별 현황", use_container_width=True)
    st.page_link("pages/guide.py", label="📖 운영 가이드", use_container_width=True)

    st.divider()
    st.markdown("**저장 데이터 내보내기**")
    _store = load_store()
    if _store:
        _dl_df = pd.DataFrame([
            {"공고번호": k, "처리상태": v["처리상태"], "메모": v["메모"], "상태변경일": v.get("상태변경일", "")}
            for k, v in _store.items()
        ])
        st.download_button(
            label=f"📥 처리내역 CSV ({len(_dl_df):,}건)",
            data=_dl_df.to_csv(index=False, encoding="utf-8-sig"),
            file_name=f"처리내역_{date.today().strftime('%Y%m%d')}.csv",
            mime="text/csv",
            use_container_width=True,
            key="dl_store",
        )
    else:
        st.caption("저장된 데이터 없음")


def validate():
    if not auth_key:
        st.error("인증키를 입력해 주세요.")
        return False
    delta = (end_date - start_date).days
    if delta < 0:
        st.error("종료일이 시작일보다 빠릅니다.")
        return False
    if delta > 30:
        st.error("최대 31일 범위까지만 조회 가능합니다.")
        return False
    return True


def fetch(start, end, auth, wanted_no):
    params = {
        "authKey":    auth,
        "returnType": "XML",
        "callTp":     "D",
        "ifDtmStdt":  start.strftime("%Y%m%d"),
        "ifDtmEndt":  end.strftime("%Y%m%d"),
    }
    if wanted_no:
        params["wantedAuthNo"] = wanted_no
    resp = requests.get(API_URL, params=params, timeout=30)
    resp.raise_for_status()
    return resp.text


def fetch_all(start, end, auth, wanted_no):
    frames = []
    errors = []
    cur = start
    while cur <= end:
        chunk_end = min(cur + timedelta(days=2), end)
        try:
            xml = fetch(cur, chunk_end, auth, wanted_no)
            chunk_df = parse(xml)
            frames.append(chunk_df)
        except Exception as e:
            errors.append(f"{cur.strftime('%Y-%m-%d')}~{chunk_end.strftime('%Y-%m-%d')}: {e}")
        cur = chunk_end + timedelta(days=1)
    if errors:
        st.warning("일부 구간 조회 실패:\n" + "\n".join(errors))
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def parse(xml_text):
    root = ET.fromstring(xml_text)
    rows = []
    for item in root.iter("monitoringErrInfo"):
        row = {}
        for key, label in COLUMNS:
            node = item.find(key)
            val = node.text.strip() if node is not None and node.text else ""
            if key == "wantedInfoUrl" and val:
                val = val.split("?")[0]
            row[label] = val
        rows.append(row)
    df = pd.DataFrame(rows, columns=[label for _, label in COLUMNS])
    df.insert(0, "처리상태", "미검토")
    df.insert(1, "메모", "")
    df.insert(2, "상태변경일", "")
    return df


def make_excel(df, start, end):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "모니터링결과"

    header_font  = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
    header_fill  = PatternFill("solid", start_color="1F5C8B")
    cell_font    = Font(name="맑은 고딕", size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin         = Side(style="thin", color="CCCCCC")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill     = PatternFill("solid", start_color="EAF2FB")
    green_fill   = PatternFill("solid", start_color="D9F2D9")
    yellow_fill  = PatternFill("solid", start_color="FFF9CC")
    blue_fill    = PatternFill("solid", start_color="D6E8FB")
    red_fill     = PatternFill("solid", start_color="FFD9D9")

    ncols = len(df.columns)
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    t = ws["A1"]
    t.value     = f"연계채용정보 모니터링 결과  |  {start.strftime('%Y%m%d')} ~ {end.strftime('%Y%m%d')}"
    t.font      = Font(name="맑은 고딕", bold=True, size=12, color="1F5C8B")
    t.alignment = center_align
    ws.row_dimensions[1].height = 28

    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(row=2, column=ci, value=col)
        c.font = header_font; c.fill = header_fill
        c.alignment = center_align; c.border = border
    ws.row_dimensions[2].height = 22

    long_cols = {"에러내용", "법령 맵핑 내용", "구인공고 URL"}
    for ri, (_, row) in enumerate(df.iterrows(), 3):
        status = row["처리상태"]
        if status == "이상없음":
            row_fill = green_fill
        elif status == "게재중단":
            row_fill = red_fill
        elif status == "검토중":
            row_fill = yellow_fill
        elif status == "검토완료":
            row_fill = blue_fill
        else:
            row_fill = alt_fill if ri % 2 == 1 else None
        for ci, col in enumerate(df.columns, 1):
            val = row[col]
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = cell_font; c.border = border
            c.alignment = left_align if col in long_cols else center_align
            if row_fill:
                c.fill = row_fill
        ws.row_dimensions[ri].height = 18

    col_widths = [12, 30, 18, 20, 20, 20, 20, 20, 50, 16, 40, 40]
    for i, w in enumerate(col_widths[:ncols], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── 메인 ───────────────────────────────────────
st.title("📋 연계채용정보 모니터링 결과조회")
st.caption("한국고용정보원 Work24 Open API")


if search_btn:
    if not validate():
        st.stop()

    with st.spinner("데이터 조회 중..."):
        try:
            raw_df = fetch_all(start_date, end_date, auth_key, wanted_auth_no)
        except Exception as e:
            st.error(f"오류 발생: {e}")
            st.stop()

        if raw_df.empty:
            # 첫 청크 응답 원문 확인
            try:
                chunk_end = min(start_date + timedelta(days=2), end_date)
                xml_raw = fetch(start_date, chunk_end, auth_key, wanted_auth_no)
                with st.expander("🔍 API 응답 원문 (디버그)", expanded=True):
                    st.code(xml_raw[:2000], language="xml")
            except Exception as e:
                st.error(f"API 직접 호출 실패: {e}")

    if raw_df.empty:
        st.session_state["base_df"]  = raw_df
        st.session_state["period"]   = (start_date, end_date)
        st.session_state["raw_df"]   = raw_df
        st.session_state.pop("data_editor", None)
        st.rerun()

    st.session_state["raw_df"] = raw_df.copy()

    df = raw_df.sort_values("연계일시", ascending=False).drop_duplicates(subset=["공고번호"]).reset_index(drop=True)

    if exclude_employment:
        df = df[~df["에러내용"].str.contains("고용형태", na=False)].reset_index(drop=True)

    stored = load_store()
    for idx in df.index:
        key = df.at[idx, "공고번호"]
        if key and key in stored:
            df.at[idx, "처리상태"]  = stored[key]["처리상태"]
            df.at[idx, "메모"]      = stored[key]["메모"]
            df.at[idx, "상태변경일"] = stored[key].get("상태변경일", "")

    st.session_state["base_df"] = df
    st.session_state["period"]  = (start_date, end_date)
    st.session_state.pop("data_editor", None)


if "base_df" not in st.session_state:
    st.info("← 왼쪽 사이드바에서 조회 조건을 입력하고 **조회** 버튼을 누르세요.")
    st.stop()

base_df = st.session_state["base_df"]
start_saved, end_saved = st.session_state["period"]

if base_df.empty:
    st.info("조회된 데이터가 없습니다.")
    st.stop()

# ── 요약 지표 (저장된 상태 기준) ──────────────
import plotly.graph_objects as go

total         = len(base_df)
이상없음_count = (base_df["처리상태"] == "이상없음").sum()
게재중단_count = (base_df["처리상태"] == "게재중단").sum()
검토중_count   = (base_df["처리상태"] == "검토중").sum()
검토완료_count = (base_df["처리상태"] == "검토완료").sum()
미검토_count   = total - 이상없음_count - 게재중단_count - 검토중_count - 검토완료_count

left, right = st.columns([3, 2])
with left:
    r1c1, r1c2, r1c3 = st.columns(3)
    r1c1.metric("총 건수",   f"{total:,}건")
    r1c2.metric("미검토",    f"{미검토_count:,}건")
    r1c3.metric("검토중",    f"{검토중_count:,}건")
    r2c1, r2c2, r2c3 = st.columns(3)
    r2c1.metric("검토완료",  f"{검토완료_count:,}건")
    r2c2.metric("이상없음",  f"{이상없음_count:,}건")
    r2c3.metric("게재중단",  f"{게재중단_count:,}건")

with right:
    labels = ["미검토", "검토중", "검토완료", "이상없음", "게재중단"]
    values = [미검토_count, 검토중_count, 검토완료_count, 이상없음_count, 게재중단_count]
    colors = ["#CCCCCC", "#FFD700", "#4096EE", "#5CB85C", "#D9534F"]
    fig = go.Figure(go.Pie(
        labels=labels, values=values, hole=0.55,
        marker_colors=colors,
        textinfo="percent", textfont_size=12,
        hovertemplate="%{label}: %{value}건 (%{percent})<extra></extra>",
    ))
    fig.update_layout(
        margin=dict(t=10, b=10, l=10, r=10), height=200,
        showlegend=True,
        legend=dict(orientation="v", x=1.05, y=0.5, font_size=11),
    )
    st.plotly_chart(fig, use_container_width=True)

st.divider()

# ── 결과 필터 ─────────────────────────────────
law_map_count  = base_df["법령 맵핑 내용"].str.strip().ne("").sum()
memo_count     = base_df["메모"].str.strip().ne("").sum()
law_volt_count = (base_df["법위반의심 여부"] == "Y").sum()

with st.expander("🔧 결과 필터", expanded=True):
    fc1, fc2 = st.columns(2)
    with fc1:
        status_filter = st.selectbox("처리 상태", ["전체", "미검토", "검토중", "검토완료", "이상없음", "게재중단"])
    with fc2:
        err_type_filter = st.selectbox("오류구분", ["전체", "사전필터링만 보기", "구인모니터링만 보기"])
    cb1, cb2, cb3 = st.columns(3)
    with cb1:
        law_volt_only = st.checkbox(f"법위반의심 Y만 ({law_volt_count:,}건)")
    with cb2:
        law_only      = st.checkbox(f"법령 맵핑 내용 있는것만 ({law_map_count:,}건)")
    with cb3:
        memo_only     = st.checkbox(f"메모 있는것만 ({memo_count:,}건)")

filtered = base_df.copy()
if status_filter != "전체":
    filtered = filtered[filtered["처리상태"] == status_filter]
if err_type_filter == "사전필터링만 보기":
    filtered = filtered[filtered["오류구분"].str.contains("사전", na=False)]
elif err_type_filter == "구인모니터링만 보기":
    filtered = filtered[filtered["오류구분"].str.contains("구인", na=False)]
if law_volt_only:
    filtered = filtered[filtered["법위반의심 여부"] == "Y"]
if law_only:
    filtered = filtered[filtered["법령 맵핑 내용"].str.strip().ne("")]
if memo_only:
    filtered = filtered[filtered["메모"].str.strip().ne("")]

filtered_display = filtered.reset_index(drop=True)
st.caption(f"필터 결과: {len(filtered_display):,}건 / 전체 {total:,}건")

# 처리상태별 색상 지시자 컬럼 (read-only)
_EMOJI = {"이상없음": "🟢", "게재중단": "🔴", "검토중": "🟡", "검토완료": "🔵", "미검토": "⬜"}
display_with_color = filtered_display.copy()
display_with_color.insert(0, "색상", display_with_color["처리상태"].map(_EMOJI).fillna("⬜"))

# 저장 시 행 인덱스 → 공고번호 매핑용
st.session_state["_fd"] = filtered_display

st.data_editor(
    display_with_color,
    column_config={
        "색상": st.column_config.TextColumn("", width="small"),
        "처리상태": st.column_config.SelectboxColumn(
            "처리상태", options=STATUS_OPTIONS, required=True, width="small",
        ),
        "상태변경일":     st.column_config.TextColumn("상태변경일",     width="small"),
        "메모":           st.column_config.TextColumn("메모",           width="medium"),
        "에러내용":       st.column_config.TextColumn("에러내용",       width="large"),
        "법령 맵핑 내용": st.column_config.TextColumn("법령 맵핑 내용", width="large"),
        "구인공고 URL":   st.column_config.LinkColumn("구인공고 URL",   width="large", display_text="🔗 바로가기"),
    },
    disabled=["색상", "상태변경일", "구인공고 URL"],
    hide_index=True,
    use_container_width=True,
    height=450,
    key="data_editor",
    # on_change 제거 — 편집 중 Streamlit 재실행이 없으므로 행 위치 유지
)

st.divider()

# ── 저장 ──────────────────────────────────────
def build_final_store(pending: dict):
    store = {}
    to_delete = set()
    for _, row in base_df.iterrows():
        k = row["공고번호"]
        s = row["처리상태"]
        m = str(row["메모"] or "")
        c = str(row.get("상태변경일", "") or "")
        if k and (s != "미검토" or m.strip()):
            store[k] = {"처리상태": s, "메모": m, "상태변경일": c}
    for k, v in pending.items():
        s = v["처리상태"]
        m = str(v.get("메모", "") or "")
        c = str(v.get("상태변경일", "") or "")
        if s != "미검토" or m.strip():
            store[k] = {"처리상태": s, "메모": m, "상태변경일": c}
        else:
            if k in store:
                del store[k]
            to_delete.add(k)
    return store, to_delete

has_changes = bool(st.session_state.get("data_editor", {}).get("edited_rows"))

btn_col, dl_col = st.columns([1, 2])

with btn_col:
    save_label = "💾 저장" + (" ●" if has_changes else "")
    if st.button(save_label, type="primary", use_container_width=True, key="save_btn"):
        pending = _get_pending_from_editor()
        current_store, to_delete_set = build_final_store(pending)
        try:
            save_store(current_store, to_delete_set)
        except Exception as e:
            st.error(f"저장 실패: {e}")
            st.stop()

        new_base = base_df.copy()
        for idx in new_base.index:
            k = new_base.at[idx, "공고번호"]
            if k and k in current_store:
                new_base.at[idx, "처리상태"]  = current_store[k]["처리상태"]
                new_base.at[idx, "메모"]      = current_store[k]["메모"]
                new_base.at[idx, "상태변경일"] = current_store[k].get("상태변경일", "")
            elif k:
                new_base.at[idx, "처리상태"]  = "미검토"
                new_base.at[idx, "메모"]      = ""
                new_base.at[idx, "상태변경일"] = ""
        st.session_state["base_df"] = new_base
        st.session_state.pop("data_editor", None)

        st.toast(f"저장 완료 ({len(current_store)}건)", icon="✅")
        st.rerun()

with dl_col:
    excel_buf = make_excel(base_df, start_saved, end_saved)
    filename  = f"모니터링결과_{start_saved.strftime('%Y%m%d')}-{end_saved.strftime('%Y%m%d')}.xlsx"
    st.download_button(
        label="📥 엑셀 다운로드 (전체)",
        data=excel_buf,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
